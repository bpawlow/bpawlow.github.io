"""Add centralized team-display-name settings to an existing shared workbook."""

from copy import copy
from pathlib import Path
import sys

from openpyxl import load_workbook


def update(path: Path) -> None:
    workbook = load_workbook(path)
    sheet = workbook["App Config"]
    existing = {sheet.cell(row, 1).value for row in range(2, sheet.max_row + 1)}
    rows = (
        ("TEAM_A_NAME", "Team A"),
        ("TEAM_B_NAME", "Team B"),
        ("TEAM_C_NAME", "Team C"),
    )
    template_row = min(sheet.max_row, 8)
    for key, value in rows:
        if key in existing:
            continue
        target_row = sheet.max_row + 1
        sheet.append([key, value, "Central display name used throughout the Sheet and website.", "Organizer"])
        for column in range(1, 5):
            source = sheet.cell(template_row, column)
            target = sheet.cell(target_row, column)
            target._style = copy(source._style)
            target.number_format = source.number_format
            target.alignment = copy(source.alignment)
        sheet.cell(target_row, 2).fill = copy(sheet["B8"].fill)
    workbook.save(path)


if __name__ == "__main__":
    update(Path(sys.argv[1]))
