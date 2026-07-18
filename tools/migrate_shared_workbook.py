"""Migrate an existing shared workbook to the flexible schedule/model schema."""

from copy import copy
from pathlib import Path
import sys

from openpyxl import load_workbook


MODEL_CONFIG = (
    ("STRAIGHT_VIG", 0.06, "Straight-market overround used for game lines and player props."),
    ("PARLAY_BASE_VIG", 0.08, "Base parlay margin applied after joint simulation pricing."),
    ("THREE_POINT_RATE_MIN", 0.22, "Minimum amateur pickup three-point attempt rate."),
    ("THREE_POINT_RATE_MAX", 0.55, "Maximum amateur pickup three-point attempt rate."),
    ("SCORING_USAGE_WEIGHT", 0.65, "How strongly scoring rating concentrates shot attempts."),
    ("SHOOTING_USAGE_WEIGHT", 0.18, "How strongly shooting rating concentrates shot attempts."),
    ("THREE_POINT_ATTEMPT_SHOOTING_WEIGHT", 0.04, "How strongly shooting rating changes three-point attempt mix."),
    ("THREE_POINT_ATTEMPT_USAGE_WEIGHT", 0.20, "How strongly player usage changes three-point attempt mix."),
    ("POINTS_MAKE_SKILL_SLOPE", 0.028, "Two-point make-probability sensitivity to individual skill."),
    ("THREE_POINT_MAKE_SKILL_SLOPE", 0.03, "Three-point make-probability sensitivity to shooting skill."),
    ("ASSIST_BASE_RATE", 0.40, "Base chance that a made basket receives an assist."),
    ("ASSIST_PLAYMAKING_SLOPE", 0.04, "Assist-rate increase per playmaking rating point."),
    ("ASSIST_ROLE_EXPONENT", 2.2, "How sharply high-playmaking players receive assist credit."),
    ("ASSIST_ROLE_WEIGHT", 1.4, "Strength of individual playmaking in assist-credit allocation."),
    ("OFFENSIVE_REBOUND_BASE_RATE", 0.28, "Base chance a miss becomes an offensive rebound."),
    ("REBOUND_ROLE_EXPONENT", 1.9, "How sharply high-rebounding players receive rebound credit."),
    ("REBOUND_ROLE_WEIGHT", 1.35, "Strength of individual rebounding in rebound allocation."),
)

DEPRECATED_CONFIG_KEYS = {
    "APPS_SCRIPT_URL", "AUTO_REFRESH_SECONDS",
    "REBOUND_LINE_QUANTILE", "ASSIST_LINE_QUANTILE", "THREES_LINE_QUANTILE",
    "POINTS_LINE_QUANTILE", "COMBO_LINE_QUANTILE",
    "POINTS_LINE_OFFSET", "REBOUNDS_LINE_OFFSET", "ASSISTS_LINE_OFFSET",
    "THREES_LINE_OFFSET", "PR_LINE_OFFSET", "PA_LINE_OFFSET",
    "RA_LINE_OFFSET", "PRA_LINE_OFFSET",
}


def copy_row_style(sheet, source_row: int, target_row: int, columns: int) -> None:
    for column in range(1, columns + 1):
        source = sheet.cell(source_row, column)
        target = sheet.cell(target_row, column)
        target._style = copy(source._style)
        target.alignment = copy(source.alignment)
        target.number_format = source.number_format


def ensure_sheet(workbook, title: str, headers: list[str]):
    sheet = workbook[title] if title in workbook.sheetnames else workbook.create_sheet(title)
    existing = [sheet.cell(1, column).value for column in range(1, sheet.max_column + 1)] if sheet.max_column else []
    if not existing or existing[0] is None:
        sheet.append(headers)
    else:
        for header in headers:
            if header not in existing:
                sheet.cell(1, sheet.max_column + 1).value = header
                existing.append(header)
    return sheet


def migrate(path: Path) -> None:
    workbook = load_workbook(path)

    config = workbook["App Config"]
    for row in range(config.max_row, 1, -1):
        if config.cell(row, 1).value in DEPRECATED_CONFIG_KEYS:
            config.delete_rows(row, 1)
    existing_keys = {config.cell(row, 1).value for row in range(2, config.max_row + 1)}
    for key, value, description in MODEL_CONFIG:
        if key in existing_keys:
            continue
        row = config.max_row + 1
        config.append([key, value, description, "Organizer"])
        copy_row_style(config, min(row - 1, 11), row, 4)
        config.cell(row, 2).fill = copy(config["B11"].fill)
    for key, value, _description, *legacy in MODEL_CONFIG:
        if not legacy:
            continue
        for row in range(2, config.max_row + 1):
            if config.cell(row, 1).value == key and config.cell(row, 2).value == legacy[0]:
                config.cell(row, 2).value = value
                break
    for row in range(2, config.max_row + 1):
        if config.cell(row, 1).value == "MODEL_VERSION" and (config.cell(row, 2).value or 0) < 4:
            config.cell(row, 2).value = 4
            break

    schedule = workbook["Schedule & Results"]
    headers = [schedule.cell(1, column).value for column in range(1, schedule.max_column + 1)]
    additions = [
        ("Game Type", "TOURNAMENT"),
        ("Team 1 ID", ""),
        ("Team 2 ID", ""),
        ("Bye ID", ""),
        ("Counts Toward Standings?", True),
        ("Betting Enabled?", True),
    ]
    for header, _ in additions:
        if header not in headers:
            schedule.cell(1, schedule.max_column + 1).value = header
            headers.append(header)
    columns = {header: index + 1 for index, header in enumerate(headers)}
    canonical = {
        "game-1": ("Team A", "Team B", "Team C"),
        "game-2": ("Team B", "Team C", "Team A"),
        "game-3": ("Team C", "Team A", "Team B"),
    }
    for row in range(2, schedule.max_row + 1):
        game_id = schedule.cell(row, columns["Game ID"]).value
        if not game_id:
            continue
        fallback = canonical.get(str(game_id), ("", "", ""))
        values = {
            "Game Type": schedule.cell(row, columns["Game Type"]).value or "TOURNAMENT",
            "Team 1 ID": schedule.cell(row, columns["Team 1 ID"]).value or fallback[0],
            "Team 2 ID": schedule.cell(row, columns["Team 2 ID"]).value or fallback[1],
            "Bye ID": schedule.cell(row, columns["Bye ID"]).value or fallback[2],
            "Counts Toward Standings?": schedule.cell(row, columns["Counts Toward Standings?"]).value if schedule.cell(row, columns["Counts Toward Standings?"]).value != "" else True,
            "Betting Enabled?": schedule.cell(row, columns["Betting Enabled?"]).value if schedule.cell(row, columns["Betting Enabled?"]).value != "" else True,
        }
        for header, value in values.items():
            schedule.cell(row, columns[header]).value = value

    beer_config = ensure_sheet(workbook, "Beer Olympics Config", ["Key", "Value", "Description"])
    existing_beer_config = {beer_config.cell(row, 1).value for row in range(2, beer_config.max_row + 1)}
    for item in (
        ("BEER_OLYMPICS_ENABLED", False, "Set TRUE to show Beer Olympics in the website."),
        ("BEER_MODEL_VERSION", 1, "Increment after changing Beer schedule or manual market assumptions."),
        ("BEER_STANDINGS_ENABLED", True, "Set FALSE to hide Beer wins/losses standings."),
        ("BEER_PARLAY_ENABLED", True, "Set FALSE to prevent Beer markets from entering parlays."),
    ):
        if item[0] not in existing_beer_config:
            beer_config.append(list(item))

    beer_schedule = ensure_sheet(workbook, "Beer Schedule & Results", ["Event ID", "Event #", "Event Name", "Matchup ID", "Sequence", "Team 1 ID", "Team 1", "Team 2 ID", "Team 2", "Status", "Betting Enabled?", "Betting Locked?", "Counts Toward Standings?", "Winner Team ID", "Team 1 Score/Time", "Team 2 Score/Time", "Final?", "Updated At", "Notes"])
    if beer_schedule.max_row < 2:
        events = [("beer-kayak", 1, "Kayak Race / Battle Royale"), ("beer-chug-cornhole", 2, "Beer Chug + Cornhole"), ("beer-die", 3, "Beer Die"), ("beer-spikeball", 4, "Spikeball"), ("beer-football", 5, "Beer Football")]
        pairs = [("Team A", "Team B"), ("Team B", "Team C"), ("Team C", "Team A")]
        rows = []
        sequence = 1
        for event_id, number, name in events:
            for matchup_number, (team1, team2) in enumerate(pairs, 1):
                rows.append([event_id, number, name, f"{event_id}-matchup-{matchup_number}", sequence, team1, team1, team2, team2, "UPCOMING", True, False, True, "", "", "", False, "", ""])
                sequence += 1
        for row in rows:
            beer_schedule.append(row)

    beer_moneylines = ensure_sheet(workbook, "Beer Moneylines", ["Matchup ID", "Team ID", "Team Name", "American Odds", "Betting Enabled?", "Notes"])
    if beer_moneylines.max_row < 2:
        for row in range(2, beer_schedule.max_row + 1):
            matchup_id = beer_schedule.cell(row, 4).value
            for team_column in (6, 8):
                team = beer_schedule.cell(row, team_column).value
                beer_moneylines.append([matchup_id, team, team, "", True, "Enter manual American odds."])

    ensure_sheet(workbook, "Beer Die Props", ["Prop ID", "Matchup ID", "Prop Name", "Scope", "Team ID", "Market Type", "Line", "Over American Odds", "Under American Odds", "Yes American Odds", "No American Odds", "Actual Result Value", "Winning Side", "Betting Enabled?", "Betting Locked?", "Final?", "Notes"])

    ensure_sheet(workbook, "Beer Bets", ["Bet ID", "Submitted At", "Bettor", "Stake", "Decimal Odds", "American Odds", "Potential Return", "Scenario", "Status", "Settled Return", "Profit", "Model Version", "Event ID"])
    ensure_sheet(workbook, "Beer Bet Legs", ["Bet ID", "Leg #", "Game ID", "Competition", "Kind", "Subject", "Player ID", "Prop ID", "Team", "Stat", "Side", "Line", "Label", "Leg Decimal Odds", "Grade"])
    ensure_sheet(workbook, "Beer Betting Leaderboard", ["Bettor", "Tickets", "Total Staked", "Settled Return", "Profit", "Units Available"])

    legs = workbook["Bet Legs"]
    leg_headers = [legs.cell(1, column).value for column in range(1, legs.max_column + 1)]
    for header in ("Competition", "Prop ID"):
        if header not in leg_headers:
            legs.cell(1, legs.max_column + 1).value = header
            leg_headers.append(header)

    beer_legs = workbook["Beer Bet Legs"]
    beer_leg_headers = [beer_legs.cell(1, column).value for column in range(1, beer_legs.max_column + 1)]
    for header in ("Competition", "Prop ID"):
        if header not in beer_leg_headers:
            beer_legs.cell(1, beer_legs.max_column + 1).value = header
            beer_leg_headers.append(header)

    if "Game Rosters" not in workbook.sheetnames:
        workbook.create_sheet("Game Rosters")
        roster = workbook["Game Rosters"]
        roster.append(["Game ID", "Roster Configuration", "Team ID", "Player", "Rotation Share", "Active?", "Notes"])
        for cell in roster[1]:
            cell.font = copy(schedule[1][0].font)
            cell.fill = copy(schedule[1][0].fill)
        roster.column_dimensions["A"].width = 13
        roster.column_dimensions["B"].width = 22
        roster.column_dimensions["C"].width = 13
        roster.column_dimensions["D"].width = 22
        roster.column_dimensions["E"].width = 16
        roster.column_dimensions["F"].width = 11
        roster.column_dimensions["G"].width = 42

    workbook.save(path)


if __name__ == "__main__":
    migrate(Path(sys.argv[1]))
