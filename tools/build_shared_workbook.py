from __future__ import annotations

import statistics
from copy import copy
from pathlib import Path

import openpyxl
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.datavalidation import DataValidation

ROOT = Path("/Users/bpawlow/personal")
PATH = ROOT / "bachelor_player_ratings_shared.xlsx"

NAVY = "172033"
BLUE = "2E5BFF"
LIGHT_BLUE = "DCE6FF"
YELLOW = "FFF2A8"
GREEN = "D9EAD3"
LIGHT_GREEN = "E8F5E9"
WHITE = "FFFFFF"
GRAY = "F3F5F8"
MID_GRAY = "D7DCE5"
RED = "FCE8E6"
thin = Side(style="thin", color=MID_GRAY)


def style_sheet(ws, widths: dict[str, float], freeze: str = "A2") -> None:
    ws.freeze_panes = freeze
    ws.auto_filter.ref = ws.dimensions
    ws.sheet_view.showGridLines = False
    for cell in ws[1]:
        cell.fill = PatternFill("solid", fgColor=NAVY)
        cell.font = Font(color=WHITE, bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = Border(bottom=thin)
    ws.row_dimensions[1].height = 30
    for column, width in widths.items():
        ws.column_dimensions[column].width = width
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.border = Border(bottom=thin)
            cell.alignment = Alignment(vertical="center", wrap_text=True)


def add_validation(ws, formula: str, cell_range: str) -> None:
    validation = DataValidation(type="list", formula1=formula, allow_blank=True)
    ws.add_data_validation(validation)
    validation.add(cell_range)


def new_sheet(wb, title: str, headers: list[str], widths: dict[str, float]):
    if title in wb.sheetnames:
        del wb[title]
    ws = wb.create_sheet(title)
    ws.append(headers)
    style_sheet(ws, widths)
    return ws


wb = openpyxl.load_workbook(PATH)
wb.calculation.fullCalcOnLoad = True
wb.calculation.forceFullCalc = True
wb.calculation.calcMode = "auto"

ratings = wb["Quick Player Ratings"]
player_rows = [row for row in range(3, ratings.max_row + 1) if ratings.cell(row, 1).value]
skill_columns = list(range(5, 11))
original: dict[tuple[int, int], float] = {}
category_means: dict[int, float] = {}

for column in skill_columns:
    values = [float(ratings.cell(row, column).value) for row in player_rows]
    category_means[column] = statistics.mean(values)
    for row in player_rows:
        old = float(ratings.cell(row, column).value)
        original[(row, column)] = old
        centered = old - category_means[column] + 5
        normalized = max(1, min(10, round(centered * 2) / 2))
        ratings.cell(row, column).value = normalized
        ratings.cell(row, column).fill = PatternFill("solid", fgColor=YELLOW)

# Make the key calculated fields Google Sheets-safe and explicit.
for row in player_rows:
    ratings.cell(row, 4).value = f"=AVERAGE(E{row}:J{row})"
    ratings.cell(row, 13).value = f'=IFERROR((D{row}*0.35)+(E{row}*0.2)+(F{row}*0.1)+(G{row}*0.1)+(H{row}*0.15)+(I{row}*0.07)+(J{row}*0.03),"")'
    ratings.cell(row, 14).value = f'=IFERROR((E{row}*0.45)+(F{row}*0.25)+(G{row}*0.2)+(D{row}*0.1),"")'
    ratings.cell(row, 15).value = f'=IFERROR((H{row}*0.6)+(I{row}*0.25)+(J{row}*0.15),"")'
    ratings.cell(row, 16).value = f'=IFERROR((E{row}+G{row})/20,"")'
    ratings.cell(row, 17).value = f'=IFERROR(IF(K{row}="High",0.8,IF(K{row}="Medium",1,IF(K{row}="Low",1.25,1))),"")'

ratings["A1"] = "NORMALIZED RATINGS — yellow skill cells are centered so this group's average is 5."
ratings["A1"].font = Font(color=WHITE, bold=True, size=12)
ratings["A1"].fill = PatternFill("solid", fgColor=BLUE)

audit = new_sheet(
    wb,
    "Rating Normalization",
    ["Player", "Category", "Original", "Category Original Mean", "Normalized", "Change", "Method"],
    {"A": 18, "B": 24, "C": 12, "D": 22, "E": 13, "F": 11, "G": 55},
)
for row in player_rows:
    for column in skill_columns:
        old = original[(row, column)]
        new = ratings.cell(row, column).value
        audit.append([
            ratings.cell(row, 1).value,
            ratings.cell(2, column).value,
            old,
            round(category_means[column], 3),
            new,
            new - old,
            "Original relative gaps preserved; category mean shifted to 5 and rounded to 0.5.",
        ])
audit.conditional_formatting.add(f"E2:E{audit.max_row}", ColorScaleRule(start_type="min", start_color="F8696B", mid_type="percentile", mid_value=50, mid_color="FFEB84", end_type="max", end_color="63BE7B"))

config = new_sheet(wb, "App Config", ["Key", "Value", "Description", "Who edits this"], {"A": 25, "B": 42, "C": 70, "D": 24})
config_rows = [
    ("BRAD_PLAYS", False, "FALSE uses the default roster; TRUE uses the alternate roster.", "Organizer"),
    ("BETTING_OPEN", True, "Set FALSE to stop new centralized bets.", "Organizer"),
    ("EVENT_ID", "bachelor-basketball-2026", "Stable identifier stored on every ticket.", "Leave unchanged"),
    ("MODEL_VERSION", 4, "Increment after materially changing ratings or pricing assumptions.", "Organizer"),
    ("STARTING_UNITS", 100, "Starting bankroll for each participant.", "Leave unchanged"),
    ("TEAM_A_NAME", "Team A", "Central display name used throughout the Sheet and website.", "Organizer"),
    ("TEAM_B_NAME", "Team B", "Central display name used throughout the Sheet and website.", "Organizer"),
    ("TEAM_C_NAME", "Team C", "Central display name used throughout the Sheet and website.", "Organizer"),
    ("STRAIGHT_VIG", 0.06, "Straight-market overround used for game lines and player props.", "Organizer"),
    ("PARLAY_BASE_VIG", 0.08, "Base parlay margin applied after joint simulation pricing.", "Organizer"),
    ("THREE_POINT_RATE_MIN", 0.22, "Minimum amateur pickup three-point attempt rate.", "Organizer"),
    ("THREE_POINT_RATE_MAX", 0.55, "Maximum amateur pickup three-point attempt rate.", "Organizer"),
    ("SCORING_USAGE_WEIGHT", 0.65, "How strongly scoring rating concentrates shot attempts.", "Organizer"),
    ("SHOOTING_USAGE_WEIGHT", 0.18, "How strongly shooting rating concentrates shot attempts.", "Organizer"),
    ("THREE_POINT_ATTEMPT_SHOOTING_WEIGHT", 0.04, "How strongly shooting rating changes three-point attempt mix.", "Organizer"),
    ("THREE_POINT_ATTEMPT_USAGE_WEIGHT", 0.20, "How strongly player usage changes three-point attempt mix.", "Organizer"),
    ("POINTS_MAKE_SKILL_SLOPE", 0.028, "Two-point make-probability sensitivity to individual skill.", "Organizer"),
    ("THREE_POINT_MAKE_SKILL_SLOPE", 0.03, "Three-point make-probability sensitivity to shooting skill.", "Organizer"),
    ("ASSIST_BASE_RATE", 0.40, "Base chance that a made basket receives an assist.", "Organizer"),
    ("ASSIST_PLAYMAKING_SLOPE", 0.04, "Assist-rate increase per playmaking rating point.", "Organizer"),
    ("ASSIST_ROLE_EXPONENT", 2.2, "How sharply high-playmaking players receive assist credit.", "Organizer"),
    ("ASSIST_ROLE_WEIGHT", 1.4, "Strength of individual playmaking in assist-credit allocation.", "Organizer"),
    ("OFFENSIVE_REBOUND_BASE_RATE", 0.28, "Base chance a miss becomes an offensive rebound.", "Organizer"),
    ("REBOUND_ROLE_EXPONENT", 1.9, "How sharply high-rebounding players receive rebound credit.", "Organizer"),
    ("REBOUND_ROLE_WEIGHT", 1.35, "Strength of individual rebounding in rebound allocation.", "Organizer"),
]
for item in config_rows:
    config.append(item)
for cell in (config.cell(row, 2) for row in range(2, config.max_row + 1) if config.cell(row, 4).value in ("Organizer", "Organizer setup")):
    cell.fill = PatternFill("solid", fgColor=YELLOW)
add_validation(config, '"TRUE,FALSE"', "B2:B3")

schedule = new_sheet(
    wb,
    "Schedule & Results",
    ["Game ID", "Game #", "Game Type", "Team 1 ID", "Team 1", "Team 2 ID", "Team 2", "Bye ID", "Bye", "Counts Toward Standings?", "Betting Enabled?", "Status", "Team 1 Score", "Team 2 Score", "Final?", "Betting Locked?", "Updated At", "Team 1 Box Points", "Team 2 Box Points", "Reconciliation"],
    {"A": 13, "B": 10, "C": 16, "D": 13, "E": 16, "F": 13, "G": 16, "H": 13, "I": 16, "J": 23, "K": 18, "L": 14, "M": 16, "N": 16, "O": 11, "P": 18, "Q": 24, "R": 20, "S": 20, "T": 24},
)
games = [
    ("game-1", 1, "Team A", "Team B", "Team C", "TOURNAMENT", True, True),
    ("game-2", 2, "Team B", "Team C", "Team A", "TOURNAMENT", True, True),
    ("game-3", 3, "Team C", "Team A", "Team B", "TOURNAMENT", True, True),
]
for game_id, number, team1, team2, bye, game_type, counts, betting in games:
    schedule.append([game_id, number, game_type, team1, team1, team2, team2, bye, bye, counts, betting, "UPCOMING", "", "", False, False, "", "", "", ""])
for row in range(2, 5):
    for column in range(10, 18):
        schedule.cell(row, column).fill = PatternFill("solid", fgColor=YELLOW)
add_validation(schedule, '"TOURNAMENT,CHAMPIONSHIP,EXHIBITION"', "C2:C100")
add_validation(schedule, '"UPCOMING,LIVE,FINAL"', "L2:L100")
add_validation(schedule, '"TRUE,FALSE"', "J2:K100")
add_validation(schedule, '"TRUE,FALSE"', "O2:P100")

beer_config = new_sheet(wb, "Beer Olympics Config", ["Key", "Value", "Description"], {"A": 28, "B": 24, "C": 72})
for item in [
    ("BEER_OLYMPICS_ENABLED", False, "Set TRUE to show Beer Olympics in the website."),
    ("BEER_MODEL_VERSION", 1, "Increment after changing Beer schedule or manual market assumptions."),
    ("BEER_STANDINGS_ENABLED", True, "Set FALSE to hide Beer wins/losses standings."),
    ("BEER_PARLAY_ENABLED", True, "Set FALSE to prevent Beer markets from entering parlays."),
]:
    beer_config.append(item)
    beer_config.cell(beer_config.max_row, 2).fill = PatternFill("solid", fgColor=YELLOW)

beer_schedule = new_sheet(
    wb,
    "Beer Schedule & Results",
    ["Event ID", "Event #", "Event Name", "Matchup ID", "Sequence", "Team 1 ID", "Team 1", "Team 2 ID", "Team 2", "Status", "Betting Enabled?", "Betting Locked?", "Counts Toward Standings?", "Winner Team ID", "Team 1 Score/Time", "Team 2 Score/Time", "Final?", "Updated At", "Notes"],
    {"A": 24, "B": 10, "C": 30, "D": 28, "E": 10, "F": 14, "G": 18, "H": 14, "I": 18, "J": 14, "K": 18, "L": 18, "M": 25, "N": 18, "O": 20, "P": 20, "Q": 10, "R": 22, "S": 38},
)
beer_events = [
    ("beer-kayak", 1, "Kayak Race / Battle Royale"),
    ("beer-chug-cornhole", 2, "Beer Chug + Cornhole"),
    ("beer-die", 3, "Beer Die"),
    ("beer-spikeball", 4, "Spikeball"),
    ("beer-football", 5, "Beer Football"),
]
beer_pairs = [("Team A", "Team B"), ("Team B", "Team C"), ("Team C", "Team A")]
beer_schedule_rows = []
sequence = 1
for event_id, event_number, event_name in beer_events:
    for matchup_number, (team1, team2) in enumerate(beer_pairs, 1):
        beer_schedule_rows.append([event_id, event_number, event_name, f"{event_id}-matchup-{matchup_number}", sequence, team1, team1, team2, team2, "UPCOMING", True, False, True, "", "", "", False, "", ""])
        sequence += 1
for row in beer_schedule_rows:
    beer_schedule.append(row)
for row in range(2, beer_schedule.max_row + 1):
    for column in range(10, 18):
        beer_schedule.cell(row, column).fill = PatternFill("solid", fgColor=YELLOW)
add_validation(beer_schedule, '"UPCOMING,LIVE,FINAL"', "J2:J100")
add_validation(beer_schedule, '"TRUE,FALSE"', "K2:M100")
add_validation(beer_schedule, '"TRUE,FALSE"', "Q2:Q100")
add_validation(beer_schedule, '"Team A,Team B,Team C"', "N2:N100")

beer_moneylines = new_sheet(wb, "Beer Moneylines", ["Matchup ID", "Team ID", "Team Name", "American Odds", "Betting Enabled?", "Notes"], {"A": 30, "B": 14, "C": 22, "D": 18, "E": 18, "F": 42})
for row in beer_schedule_rows:
    for team in (row[5], row[7]):
        beer_moneylines.append([row[3], team, team, "", True, "Enter manual American odds."])
for row in range(2, beer_moneylines.max_row + 1):
    beer_moneylines.cell(row, 4).fill = PatternFill("solid", fgColor=YELLOW)
add_validation(beer_moneylines, '"TRUE,FALSE"', "E2:E200")

beer_props = new_sheet(wb, "Beer Die Props", ["Prop ID", "Matchup ID", "Prop Name", "Scope", "Team ID", "Market Type", "Line", "Over American Odds", "Under American Odds", "Yes American Odds", "No American Odds", "Actual Result Value", "Winning Side", "Betting Enabled?", "Betting Locked?", "Final?", "Notes"], {"A": 24, "B": 30, "C": 28, "D": 14, "E": 14, "F": 16, "G": 10, "H": 18, "I": 20, "J": 18, "K": 18, "L": 20, "M": 16, "N": 18, "O": 18, "P": 10, "Q": 42})
for row in range(2, 102):
    for column in range(1, 18):
        beer_props.cell(row, column).fill = PatternFill("solid", fgColor=YELLOW)
add_validation(beer_props, '"team,matchup"', "D2:D100")
add_validation(beer_props, '"yes-no,over-under"', "F2:F100")
add_validation(beer_props, '"over,under,yes,no"', "M2:M100")
add_validation(beer_props, '"TRUE,FALSE"', "N2:P100")

assignments = wb["Team Assignments"]
assignment_rows = []
for row in range(2, assignments.max_row + 1):
    scenario = assignments.cell(row, 1).value
    team = assignments.cell(row, 2).value
    name = assignments.cell(row, 4).value
    if not scenario or not name:
        continue
    names = ["Berler", "Jason"] if name == "Berler/Jason" else [name]
    for expanded_name in names:
        assignment_rows.append((scenario, team, expanded_name, 0.5 if name == "Berler/Jason" else 1.0))

game_rosters = new_sheet(wb, "Game Rosters", ["Game ID", "Roster Configuration", "Team ID", "Player", "Rotation Share", "Active?", "Notes"], {"A": 13, "B": 22, "C": 13, "D": 22, "E": 16, "F": 11, "G": 42})
add_validation(game_rosters, '"DEFAULT,ALTERNATE"', "B2:B100")
add_validation(game_rosters, '"TRUE,FALSE"', "F2:F100")

box = new_sheet(
    wb,
    "Box Scores",
    ["Game ID", "Scenario", "Player ID", "Player", "Team", "Played?", "Points", "Rebounds", "Assists", "Three Pointers", "PRA", "Scorekeeper Notes"],
    {"A": 13, "B": 15, "C": 18, "D": 18, "E": 13, "F": 11, "G": 11, "H": 12, "I": 11, "J": 16, "K": 11, "L": 38},
)
for scenario in ("Brad Out", "Brad Plays"):
    teams_by_game = {game_id: {team1, team2} for game_id, _, team1, team2, _, _, _, _ in games}
    for game_id, team_ids in teams_by_game.items():
        for row_scenario, team, player, share in assignment_rows:
            if row_scenario != scenario or team not in team_ids:
                continue
            player_id = player.lower().replace(" ", "-")
            box.append([game_id, scenario, player_id, player, team, False, "", "", "", "", f'=IF(F{box.max_row + 1},SUM(G{box.max_row + 1}:I{box.max_row + 1}),"")', "Shared sub slot" if share < 1 else ""])
for row in range(2, box.max_row + 1):
    for column in range(6, 11):
        box.cell(row, column).fill = PatternFill("solid", fgColor=YELLOW)
add_validation(box, '"TRUE,FALSE"', f"F2:F{box.max_row}")

schedule_headers = {schedule.cell(1, column).value: column for column in range(1, schedule.max_column + 1)}
for row in range(2, 5):
    scenario_formula = 'IF(\'App Config\'!$B$2,"Brad Plays","Brad Out")'
    score_columns = {"Team 1 Box Points": f'=SUMIFS(\'Box Scores\'!$G:$G,\'Box Scores\'!$A:$A,$A{row},\'Box Scores\'!$B:$B,{scenario_formula},\'Box Scores\'!$E:$E,$D{row},\'Box Scores\'!$F:$F,TRUE)', "Team 2 Box Points": f'=SUMIFS(\'Box Scores\'!$G:$G,\'Box Scores\'!$A:$A,$A{row},\'Box Scores\'!$B:$B,{scenario_formula},\'Box Scores\'!$E:$E,$F{row},\'Box Scores\'!$F:$F,TRUE)', "Reconciliation": f'=IF(NOT($O{row}),"PENDING",IF(AND($M{row}=$R{row},$N{row}=$S{row}),"OK","CHECK PLAYER POINTS"))'}
    for header, formula in score_columns.items():
        schedule.cell(row, schedule_headers[header]).value = formula
    for column in range(schedule_headers["Team 1 Box Points"], schedule_headers["Reconciliation"] + 1):
        schedule.cell(row, column).fill = PatternFill("solid", fgColor=GREEN)

participants = new_sheet(wb, "Participants", ["Bettor", "Active?", "Starting Units", "Notes"], {"A": 24, "B": 12, "C": 18, "D": 48})
for row in range(2, 22):
    participants.append(["", True, 100, ""])
    for column in range(1, 5):
        participants.cell(row, column).fill = PatternFill("solid", fgColor=YELLOW)
add_validation(participants, '"TRUE,FALSE"', "B2:B21")

bets = new_sheet(
    wb,
    "Bets",
    ["Bet ID", "Submitted At", "Bettor", "Stake", "Decimal Odds", "American Odds", "Potential Return", "Scenario", "Status", "Settled Return", "Profit", "Model Version", "Event ID"],
    {"A": 38, "B": 24, "C": 22, "D": 12, "E": 15, "F": 15, "G": 18, "H": 15, "I": 13, "J": 18, "K": 13, "L": 15, "M": 28},
)

legs = new_sheet(
    wb,
    "Bet Legs",
    ["Bet ID", "Leg #", "Game ID", "Competition", "Kind", "Subject", "Player ID", "Prop ID", "Team", "Stat", "Side", "Line", "Label", "Leg Decimal Odds", "Grade"],
    {"A": 38, "B": 10, "C": 28, "D": 18, "E": 16, "F": 28, "G": 18, "H": 24, "I": 13, "J": 13, "K": 12, "L": 11, "M": 45, "N": 20, "O": 12},
)

bet_board = new_sheet(wb, "Betting Leaderboard", ["Bettor", "Tickets", "Total Staked", "Settled Return", "Profit", "Units Available"], {"A": 24, "B": 12, "C": 18, "D": 20, "E": 14, "F": 18})
bet_board["A2"] = '=IFERROR(QUERY({Bets!C2:C,Bets!D2:D,Bets!J2:J,Bets!K2:K},"select Col1,count(Col1),sum(Col2),sum(Col3),sum(Col4) where Col1 is not null group by Col1 order by sum(Col4) desc label Col1 \'Bettor\',count(Col1) \'Tickets\',sum(Col2) \'Total Staked\',sum(Col3) \'Settled Return\',sum(Col4) \'Profit\'",0),"")'
bet_board["F2"] = '=ARRAYFORMULA(IF(A2:A="","",100-C2:C+D2:D))'

player_board = new_sheet(wb, "Player Leaderboard", ["Player", "Games", "Points", "Rebounds", "Assists", "Three Pointers", "PRA"], {"A": 24, "B": 12, "C": 13, "D": 14, "E": 12, "F": 18, "G": 13})
player_board["A2"] = '=IFERROR(QUERY(\'Box Scores\'!A2:K,"select D,count(D),sum(G),sum(H),sum(I),sum(J),sum(K) where F = TRUE and B = \'"&IF(\'App Config\'!B2,"Brad Plays","Brad Out")&"\' group by D order by sum(K) desc label D \'Player\',count(D) \'Games\',sum(G) \'Points\',sum(H) \'Rebounds\',sum(I) \'Assists\',sum(J) \'Three Pointers\',sum(K) \'PRA\'",0),"")'

# Replace the broken imported array formulas with static, normalized team diagnostics.
team_sheet = wb["Team Ratings"]
headers = ["Scenario", "Team", "Players", "Avg Overall", "Avg Offense", "Avg Defense", "Avg Rebounding", "Avg Stamina", "Team Rating", "Suggested Spread vs Scenario Avg"]
for col, header in enumerate(headers, 1):
    team_sheet.cell(1, col).value = header
player_values = {}
for row in player_rows:
    name = ratings.cell(row, 1).value
    vals = [float(ratings.cell(row, col).value) for col in skill_columns]
    overall = sum(vals) / len(vals)
    offense = vals[0] * .45 + vals[1] * .25 + vals[2] * .2 + overall * .1
    defense = vals[3] * .6 + vals[4] * .25 + vals[5] * .15
    player_values[name] = (overall, offense, defense, vals[4], vals[5])
diagnostics = []
for scenario in ("Brad Out", "Brad Plays"):
    scenario_rows = []
    for team in ("Team A", "Team B", "Team C"):
        members = [(name, share) for row_scenario, row_team, name, share in assignment_rows if row_scenario == scenario and row_team == team]
        denominator = sum(share for _, share in members)
        averages = [sum(player_values[name][index] * share for name, share in members) / denominator for index in range(5)]
        rating = averages[0] * .45 + averages[1] * .25 + averages[2] * .2 + averages[3] * .07 + averages[4] * .03
        scenario_rows.append([scenario, team, ", ".join(name + (" (50%)" if share < 1 else "") for name, share in members), *averages, rating])
    scenario_mean = statistics.mean(row[-1] for row in scenario_rows)
    for row in scenario_rows:
        diagnostics.append(row + [(row[-1] - scenario_mean) * 2.5])
for row_index, values in enumerate(diagnostics, 2):
    for column, value in enumerate(values, 1):
        team_sheet.cell(row_index, column).value = value
        if column >= 4:
            team_sheet.cell(row_index, column).number_format = "0.00"
style_sheet(team_sheet, {"A": 15, "B": 13, "C": 55, "D": 14, "E": 14, "F": 14, "G": 16, "H": 14, "I": 14, "J": 24})

readme = wb["READ ME FIRST"]
readme["A1"] = "SHARED APP CONTROL CENTER"
readme["A1"].font = Font(color=WHITE, bold=True, size=14)
readme["A1"].fill = PatternFill("solid", fgColor=BLUE)
instructions = [
    ("Roster mode", "Set App Config!B2 to TRUE or FALSE. The app uses this as the shared roster configuration."),
    ("Before each game", "Set Schedule & Results status and Betting Locked?; use FINAL only after scores and box scores are checked."),
    ("Official scores", "Enter team scores in Schedule & Results, then mark Final? TRUE."),
    ("Official box scores", "In Box Scores, use the active scenario rows, mark Played? TRUE, and enter PTS/REB/AST/3PM."),
    ("Bets", "Do not hand-edit Bets or Bet Legs after Apps Script setup; the app appends ticket snapshots."),
    ("Leaderboards", "Betting Leaderboard and Player Leaderboard derive from the raw Bets and Box Scores tabs."),
    ("Yellow cells", "Organizer/scorekeeper inputs. Dark headers and calculated cells should not be edited."),
]
start = 13
for index, item in enumerate(instructions, start):
    readme.cell(index, 1).value = item[0]
    readme.cell(index, 2).value = item[1]
    readme.cell(index, 1).font = Font(bold=True)
    readme.cell(index, 1).fill = PatternFill("solid", fgColor=LIGHT_BLUE)
readme.column_dimensions["A"].width = 24
readme.column_dimensions["B"].width = 100

preferred_order = [
    "READ ME FIRST", "App Config", "Quick Player Ratings", "Rating Normalization", "Team Assignments", "Game Rosters", "Team Ratings",
    "Schedule & Results", "Beer Olympics Config", "Beer Schedule & Results", "Beer Moneylines", "Beer Die Props", "Box Scores", "Participants", "Bets", "Bet Legs", "Betting Leaderboard", "Player Leaderboard",
]
wb._sheets = [wb[name] for name in preferred_order]
wb.save(PATH)
print(f"Saved {PATH}")
print("Normalized category means:")
for column in skill_columns:
    values = [ratings.cell(row, column).value for row in player_rows]
    print(ratings.cell(2, column).value, round(statistics.mean(values), 3), min(values), max(values))
